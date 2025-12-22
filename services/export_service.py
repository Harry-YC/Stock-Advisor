"""
Export Service

Generates PDF and Excel reports for stock analysis and portfolio data.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any

from config import settings

logger = logging.getLogger(__name__)


class ExportService:
    """
    Service for exporting analysis and portfolio data.

    Supports:
    - PDF reports with charts and expert analysis
    - Excel spreadsheets with portfolio data
    """

    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = Path(output_dir or settings.EXPORTS_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_analysis_pdf(
        self,
        symbol: str,
        stock_data: Dict[str, Any],
        expert_responses: Dict[str, str],
        chart_path: Optional[str] = None
    ) -> str:
        """
        Generate PDF report with expert analysis.

        Args:
            symbol: Stock ticker
            stock_data: Dict with quote, financials, etc.
            expert_responses: Dict mapping expert names to responses
            chart_path: Optional path to chart image

        Returns:
            Path to generated PDF
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer,
                Table, TableStyle, Image, PageBreak
            )
            from reportlab.lib import colors
        except ImportError:
            logger.error("reportlab not installed. Run: pip install reportlab")
            return ""

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_analysis_{timestamp}.pdf"
        filepath = self.output_dir / filename

        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
        )
        body_style = ParagraphStyle(
            'CustomBody',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
        )

        story = []

        # Title
        story.append(Paragraph(f"{symbol} Stock Analysis Report", title_style))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            body_style
        ))
        story.append(Spacer(1, 20))

        # Chart
        if chart_path and Path(chart_path).exists():
            try:
                img = Image(chart_path, width=6*inch, height=4*inch)
                story.append(img)
                story.append(Spacer(1, 20))
            except Exception as e:
                logger.warning(f"Failed to add chart to PDF: {e}")

        # Key Metrics Table
        if stock_data:
            story.append(Paragraph("Key Metrics", heading_style))

            quote = stock_data.get("quote", {})
            financials = stock_data.get("financials", {})

            metrics_data = [
                ["Metric", "Value"],
                ["Current Price", f"${quote.get('current_price', 'N/A')}"],
                ["Change", f"{quote.get('change_percent', 0):.2f}%"],
                ["52W High", f"${quote.get('high_52w', 'N/A')}"],
                ["52W Low", f"${quote.get('low_52w', 'N/A')}"],
                ["Market Cap", self._format_large_number(financials.get('market_cap'))],
                ["P/E Ratio", f"{financials.get('pe_ratio', 'N/A')}"],
                ["EPS", f"${financials.get('eps', 'N/A')}"],
            ]

            table = Table(metrics_data, colWidths=[2*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(table)
            story.append(Spacer(1, 20))

        # Expert Opinions
        if expert_responses:
            story.append(PageBreak())
            story.append(Paragraph("Expert Analysis", title_style))

            for expert_name, response in expert_responses.items():
                story.append(Paragraph(f"{expert_name}", heading_style))

                # Clean and truncate response for PDF
                clean_response = self._clean_for_pdf(response)
                for para in clean_response.split('\n\n'):
                    if para.strip():
                        story.append(Paragraph(para.strip(), body_style))
                        story.append(Spacer(1, 6))

                story.append(Spacer(1, 15))

        # Footer
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            "This report is for informational purposes only and does not constitute financial advice.",
            ParagraphStyle('Disclaimer', parent=body_style, fontSize=8, textColor=colors.grey)
        ))

        try:
            doc.build(story)
            logger.info(f"PDF report saved to {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"Failed to build PDF: {e}")
            return ""

    def export_portfolio_excel(
        self,
        positions: List[Dict[str, Any]],
        trades: Optional[List[Dict[str, Any]]] = None,
        snapshots: Optional[List[Dict[str, Any]]] = None
    ) -> str:
        """
        Generate Excel spreadsheet with portfolio data.

        Args:
            positions: List of position dicts
            trades: Optional list of trade dicts
            snapshots: Optional list of snapshot dicts

        Returns:
            Path to generated Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return ""

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"portfolio_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        wb = openpyxl.Workbook()

        # Positions Sheet
        ws = wb.active
        ws.title = "Positions"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")

        # Position headers
        headers = ["Symbol", "Shares", "Avg Price", "Cost Basis", "Current Price", "Value", "P&L $", "P&L %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Position data
        for row, pos in enumerate(positions, 2):
            ws.cell(row=row, column=1, value=pos.get("symbol", ""))
            ws.cell(row=row, column=2, value=pos.get("shares", 0))
            ws.cell(row=row, column=3, value=pos.get("avg_price", 0))
            ws.cell(row=row, column=4, value=pos.get("cost_basis", 0))
            ws.cell(row=row, column=5, value=pos.get("current_price", 0))
            ws.cell(row=row, column=6, value=pos.get("value", 0))
            ws.cell(row=row, column=7, value=pos.get("pnl_dollars", 0))
            ws.cell(row=row, column=8, value=pos.get("pnl_percent", 0))

        # Auto-width columns
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Trades Sheet
        if trades:
            ws2 = wb.create_sheet("Trade History")

            trade_headers = ["Date", "Symbol", "Action", "Shares", "Price", "Fees", "Total", "Notes"]
            for col, header in enumerate(trade_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row, trade in enumerate(trades, 2):
                ws2.cell(row=row, column=1, value=trade.get("trade_date", ""))
                ws2.cell(row=row, column=2, value=trade.get("symbol", ""))
                ws2.cell(row=row, column=3, value=trade.get("action", ""))
                ws2.cell(row=row, column=4, value=trade.get("shares", 0))
                ws2.cell(row=row, column=5, value=trade.get("price", 0))
                ws2.cell(row=row, column=6, value=trade.get("fees", 0))
                ws2.cell(row=row, column=7, value=trade.get("total_value", 0))
                ws2.cell(row=row, column=8, value=trade.get("notes", ""))

            for col in range(1, 9):
                ws2.column_dimensions[get_column_letter(col)].width = 15

        # Performance Sheet
        if snapshots:
            ws3 = wb.create_sheet("Performance")

            perf_headers = ["Date", "Total Value", "Total Cost", "P&L $", "P&L %"]
            for col, header in enumerate(perf_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row, snap in enumerate(snapshots, 2):
                ws3.cell(row=row, column=1, value=snap.get("date", ""))
                ws3.cell(row=row, column=2, value=snap.get("total_value", 0))
                ws3.cell(row=row, column=3, value=snap.get("total_cost", 0))
                ws3.cell(row=row, column=4, value=snap.get("daily_pnl", 0))
                ws3.cell(row=row, column=5, value=snap.get("daily_pnl_pct", 0))

            for col in range(1, 6):
                ws3.column_dimensions[get_column_letter(col)].width = 15

        try:
            wb.save(filepath)
            logger.info(f"Excel report saved to {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"Failed to save Excel: {e}")
            return ""

    def export_watchlist_csv(self, watchlist: List[Dict[str, Any]]) -> str:
        """
        Export watchlist to CSV.

        Args:
            watchlist: List of watchlist items

        Returns:
            Path to CSV file
        """
        import csv

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"watchlist_{timestamp}.csv"
        filepath = self.output_dir / filename

        try:
            with open(filepath, 'w', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=["symbol", "added_at", "notes", "tags"])
                writer.writeheader()
                for item in watchlist:
                    writer.writerow({
                        "symbol": item.get("symbol", ""),
                        "added_at": item.get("added_at", ""),
                        "notes": item.get("notes", ""),
                        "tags": ",".join(item.get("tags", [])),
                    })
            logger.info(f"CSV exported to {filepath}")
            return str(filepath)
        except Exception as e:
            logger.error(f"Failed to export CSV: {e}")
            return ""

    def _clean_for_pdf(self, text: str, max_length: int = 3000) -> str:
        """Clean text for PDF rendering."""
        # Remove markdown formatting that doesn't render well
        import re
        text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)  # Bold
        text = re.sub(r'\*([^*]+)\*', r'\1', text)  # Italic
        text = re.sub(r'#+\s*', '', text)  # Headers
        text = re.sub(r'`([^`]+)`', r'\1', text)  # Code
        text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', text)  # Links

        # Truncate if too long
        if len(text) > max_length:
            text = text[:max_length] + "..."

        return text

    def _format_large_number(self, value: Any) -> str:
        """Format large numbers with B/M suffixes."""
        if value is None:
            return "N/A"
        try:
            num = float(value)
            if num >= 1e12:
                return f"${num/1e12:.2f}T"
            elif num >= 1e9:
                return f"${num/1e9:.2f}B"
            elif num >= 1e6:
                return f"${num/1e6:.2f}M"
            else:
                return f"${num:,.0f}"
        except (ValueError, TypeError):
            return str(value)


# Convenience function
def get_export_service() -> ExportService:
    """Get export service instance."""
    return ExportService()
