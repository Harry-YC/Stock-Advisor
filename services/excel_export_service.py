"""
Travel Itinerary Excel Export Service

Exports travel plans to Excel with multiple sheets:
- Overview: Trip summary (destination, dates, travelers, budget)
- Itinerary: Day-by-day activities and schedule
- Accommodations: Hotel/rental recommendations
- Transportation: Flights, trains, transfers
- Budget: Cost breakdown by category
- Expert Tips: Recommendations from each expert
"""

import re
import logging
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, field

logger = logging.getLogger(__name__)


@dataclass
class TripOverview:
    """Trip overview data."""
    destination: str = ""
    start_date: str = ""
    end_date: str = ""
    duration_days: int = 0
    travelers: int = 1
    total_budget: float = 0.0
    currency: str = "USD"
    trip_type: str = ""  # e.g., "Leisure", "Adventure", "Cultural"


@dataclass
class DayActivity:
    """Single activity in the itinerary."""
    day: int
    date: str
    time: str
    activity: str
    location: str
    duration: str
    cost: float = 0.0
    notes: str = ""


@dataclass
class Accommodation:
    """Accommodation recommendation."""
    name: str
    type: str  # Hotel, Hostel, Airbnb, etc.
    location: str
    check_in: str
    check_out: str
    price_per_night: float
    total_price: float
    rating: Optional[float] = None
    amenities: str = ""
    booking_link: str = ""
    notes: str = ""


@dataclass
class Transportation:
    """Transportation segment."""
    type: str  # Flight, Train, Bus, Taxi, etc.
    from_location: str
    to_location: str
    date: str
    departure_time: str
    arrival_time: str
    duration: str
    carrier: str = ""
    price: float = 0.0
    booking_info: str = ""
    notes: str = ""


@dataclass
class BudgetItem:
    """Budget line item."""
    category: str
    item: str
    estimated_cost: float
    currency: str = "USD"
    notes: str = ""


@dataclass
class ExpertTip:
    """Tip from a travel expert."""
    expert: str
    category: str
    tip: str
    priority: str = "Medium"  # High, Medium, Low


@dataclass
class TravelPlan:
    """Complete travel plan for export."""
    overview: TripOverview = field(default_factory=TripOverview)
    itinerary: List[DayActivity] = field(default_factory=list)
    accommodations: List[Accommodation] = field(default_factory=list)
    transportation: List[Transportation] = field(default_factory=list)
    budget: List[BudgetItem] = field(default_factory=list)
    expert_tips: List[ExpertTip] = field(default_factory=list)
    raw_question: str = ""
    raw_recommendation: str = ""


class TravelPlanParser:
    """
    Parses expert responses to extract structured travel plan data.

    Uses regex patterns to identify common travel plan elements
    from free-form expert responses.
    """

    def __init__(self):
        # Patterns for extracting structured data
        self.day_pattern = re.compile(r"day\s*(\d+)[:\s-]+(.+?)(?=day\s*\d+|$)", re.IGNORECASE | re.DOTALL)
        self.time_pattern = re.compile(r"(\d{1,2}:\d{2}\s*(?:am|pm)?|\d{1,2}\s*(?:am|pm))", re.IGNORECASE)
        self.price_pattern = re.compile(r"\$\s*([\d,]+(?:\.\d{2})?)|(\d+(?:,\d{3})*)\s*(?:USD|EUR|GBP|JPY)")
        self.hotel_pattern = re.compile(r"(?:hotel|stay|accommodation|lodge|hostel|airbnb)[:\s]+([^\n]+)", re.IGNORECASE)
        self.flight_pattern = re.compile(r"(?:flight|fly|airline)[:\s]+([^\n]+)", re.IGNORECASE)
        self.duration_pattern = re.compile(r"(\d+)\s*(?:day|night|week)s?", re.IGNORECASE)
        self.budget_pattern = re.compile(r"(?:budget|cost|price|total)[:\s]*\$?\s*([\d,]+(?:\.\d{2})?)", re.IGNORECASE)

    def parse_expert_responses(
        self,
        question: str,
        recommendation: str,
        expert_responses: Dict[str, Dict[str, Any]]
    ) -> TravelPlan:
        """
        Parse expert responses into a structured TravelPlan.

        Args:
            question: Original travel question
            recommendation: Final synthesized recommendation
            expert_responses: Dict of expert name -> response dict

        Returns:
            TravelPlan with extracted data
        """
        plan = TravelPlan(
            raw_question=question,
            raw_recommendation=recommendation
        )

        # Parse overview from question
        plan.overview = self._parse_overview(question, recommendation)

        # Parse itinerary from recommendation
        plan.itinerary = self._parse_itinerary(recommendation)

        # Parse accommodations from relevant experts
        plan.accommodations = self._parse_accommodations(expert_responses)

        # Parse transportation
        plan.transportation = self._parse_transportation(expert_responses)

        # Parse budget
        plan.budget = self._parse_budget(recommendation, expert_responses)

        # Extract expert tips
        plan.expert_tips = self._extract_expert_tips(expert_responses)

        return plan

    def _parse_overview(self, question: str, recommendation: str) -> TripOverview:
        """Extract trip overview from question and recommendation."""
        overview = TripOverview()

        combined_text = f"{question} {recommendation}".lower()

        # Extract destination (look for common patterns)
        dest_patterns = [
            r"trip to\s+([A-Za-z\s,]+?)(?:\s+in|\s+for|\s+with|\?|$)",
            r"visiting\s+([A-Za-z\s,]+?)(?:\s+in|\s+for|\s+with|\?|$)",
            r"travel to\s+([A-Za-z\s,]+?)(?:\s+in|\s+for|\s+with|\?|$)",
        ]
        for pattern in dest_patterns:
            match = re.search(pattern, question, re.IGNORECASE)
            if match:
                overview.destination = match.group(1).strip().title()
                break

        # Extract duration
        duration_match = self.duration_pattern.search(combined_text)
        if duration_match:
            overview.duration_days = int(duration_match.group(1))

        # Extract budget
        budget_match = self.budget_pattern.search(combined_text)
        if budget_match:
            budget_str = budget_match.group(1).replace(",", "")
            try:
                overview.total_budget = float(budget_str)
            except ValueError:
                pass

        # Extract travelers
        travelers_match = re.search(r"(\d+)\s*(?:people|person|travelers|adults)", combined_text)
        if travelers_match:
            overview.travelers = int(travelers_match.group(1))

        return overview

    def _parse_itinerary(self, recommendation: str) -> List[DayActivity]:
        """Extract day-by-day itinerary from recommendation."""
        activities = []

        # Find day-by-day sections
        day_matches = self.day_pattern.findall(recommendation)

        for day_num, content in day_matches:
            try:
                day = int(day_num)
            except ValueError:
                continue

            # Split content into potential activities
            lines = [l.strip() for l in content.split('\n') if l.strip()]

            for line in lines:
                # Skip header/summary lines
                if len(line) < 10:
                    continue

                # Extract time if present
                time_match = self.time_pattern.search(line)
                time_str = time_match.group(1) if time_match else ""

                # Extract cost if present
                cost = 0.0
                price_match = self.price_pattern.search(line)
                if price_match:
                    try:
                        cost = float((price_match.group(1) or price_match.group(2)).replace(",", ""))
                    except (ValueError, AttributeError):
                        pass

                activity = DayActivity(
                    day=day,
                    date="",
                    time=time_str,
                    activity=line[:200],  # Truncate long descriptions
                    location="",
                    duration="",
                    cost=cost
                )
                activities.append(activity)

        return activities

    def _parse_accommodations(self, expert_responses: Dict) -> List[Accommodation]:
        """Extract accommodations from expert responses."""
        accommodations = []

        # Look in Accommodation Specialist responses
        for expert, response in expert_responses.items():
            content = response.get('content', '')

            # Find hotel mentions
            hotel_matches = self.hotel_pattern.findall(content)
            for hotel_text in hotel_matches[:5]:  # Limit to 5
                # Extract price if present
                price_match = self.price_pattern.search(hotel_text)
                price = 0.0
                if price_match:
                    try:
                        price = float((price_match.group(1) or price_match.group(2)).replace(",", ""))
                    except (ValueError, AttributeError):
                        pass

                accommodation = Accommodation(
                    name=hotel_text[:100].strip(),
                    type="Hotel",
                    location="",
                    check_in="",
                    check_out="",
                    price_per_night=price,
                    total_price=price,
                    notes=f"Recommended by {expert}"
                )
                accommodations.append(accommodation)

        return accommodations

    def _parse_transportation(self, expert_responses: Dict) -> List[Transportation]:
        """Extract transportation from expert responses."""
        transportation = []

        for expert, response in expert_responses.items():
            content = response.get('content', '')

            # Find flight mentions
            flight_matches = self.flight_pattern.findall(content)
            for flight_text in flight_matches[:5]:
                # Extract price if present
                price_match = self.price_pattern.search(flight_text)
                price = 0.0
                if price_match:
                    try:
                        price = float((price_match.group(1) or price_match.group(2)).replace(",", ""))
                    except (ValueError, AttributeError):
                        pass

                transport = Transportation(
                    type="Flight",
                    from_location="",
                    to_location="",
                    date="",
                    departure_time="",
                    arrival_time="",
                    duration="",
                    price=price,
                    notes=flight_text[:200].strip()
                )
                transportation.append(transport)

        return transportation

    def _parse_budget(
        self,
        recommendation: str,
        expert_responses: Dict
    ) -> List[BudgetItem]:
        """Extract budget items from responses."""
        budget_items = []

        # Common budget categories
        categories = {
            "Accommodation": ["hotel", "stay", "lodging", "airbnb"],
            "Transportation": ["flight", "train", "taxi", "uber", "transport"],
            "Food & Dining": ["food", "restaurant", "meal", "dining"],
            "Activities": ["tour", "ticket", "entrance", "activity"],
            "Shopping": ["shopping", "souvenir"],
            "Miscellaneous": ["misc", "other", "emergency"]
        }

        combined_text = recommendation
        for response in expert_responses.values():
            combined_text += " " + response.get('content', '')

        for category, keywords in categories.items():
            # Find price mentions near category keywords
            for keyword in keywords:
                pattern = rf"{keyword}[^$]*?\$\s*([\d,]+(?:\.\d{{2}})?)"
                matches = re.findall(pattern, combined_text, re.IGNORECASE)

                for price_str in matches[:3]:  # Limit per category
                    try:
                        cost = float(price_str.replace(",", ""))
                        item = BudgetItem(
                            category=category,
                            item=f"{keyword.title()} expense",
                            estimated_cost=cost
                        )
                        budget_items.append(item)
                        break  # One per keyword
                    except ValueError:
                        continue

        return budget_items

    def _extract_expert_tips(self, expert_responses: Dict) -> List[ExpertTip]:
        """Extract actionable tips from each expert."""
        tips = []

        # Map experts to tip categories
        expert_categories = {
            "Budget Advisor": "Budget",
            "Safety Expert": "Safety",
            "Local Culture Guide": "Culture",
            "Logistics Planner": "Logistics",
            "Food & Dining Expert": "Food",
            "Activity Curator": "Activities",
            "Accommodation Specialist": "Accommodation",
            "Weather Analyst": "Weather"
        }

        for expert, response in expert_responses.items():
            content = response.get('content', '')
            category = expert_categories.get(expert, "General")

            # Extract sentences that look like tips/recommendations
            tip_patterns = [
                r"(?:recommend|suggest|tip|advice|should|consider)[:\s]+([^.!?]+[.!?])",
                r"(?:don't forget|make sure|be sure)[:\s]+([^.!?]+[.!?])",
                r"(?:pro tip|insider tip)[:\s]+([^.!?]+[.!?])",
            ]

            for pattern in tip_patterns:
                matches = re.findall(pattern, content, re.IGNORECASE)
                for tip_text in matches[:2]:  # Max 2 tips per pattern
                    tip = ExpertTip(
                        expert=expert,
                        category=category,
                        tip=tip_text.strip()[:300],
                        priority="Medium"
                    )
                    tips.append(tip)

        return tips


class ExcelExportService:
    """
    Exports travel plans to Excel workbooks.

    Creates multi-sheet workbooks with:
    - Overview sheet with trip summary
    - Day-by-day itinerary
    - Accommodation recommendations
    - Transportation details
    - Budget breakdown
    - Expert tips and recommendations
    """

    def __init__(self):
        self.parser = TravelPlanParser()

    def export_to_excel(
        self,
        question: str,
        recommendation: str,
        expert_responses: Dict[str, Dict[str, Any]],
        filename: Optional[str] = None,
        trip_data: Optional[Dict] = None
    ) -> BytesIO:
        """
        Export travel plan to Excel workbook.

        Args:
            question: Original travel question
            recommendation: Final synthesized recommendation
            expert_responses: Dict of expert responses
            filename: Optional filename (not used, returns BytesIO)
            trip_data: Optional real-time trip data (weather, flights, hotels, dining)

        Returns:
            BytesIO buffer containing Excel workbook
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise RuntimeError("openpyxl required for Excel export. Install with: pip install openpyxl")

        # Parse expert responses into structured data
        plan = self.parser.parse_expert_responses(question, recommendation, expert_responses)

        # Create workbook
        wb = openpyxl.Workbook()

        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Overview
        ws_overview = wb.active
        ws_overview.title = "Overview"
        self._create_overview_sheet(ws_overview, plan, header_font, header_fill, title_font, thin_border)

        # Sheet 2: Itinerary
        ws_itinerary = wb.create_sheet("Itinerary")
        self._create_itinerary_sheet(ws_itinerary, plan, header_font, header_fill, thin_border)

        # Sheet 3: Accommodations
        ws_accommodations = wb.create_sheet("Accommodations")
        self._create_accommodations_sheet(ws_accommodations, plan, header_font, header_fill, thin_border)

        # Sheet 4: Transportation
        ws_transport = wb.create_sheet("Transportation")
        self._create_transportation_sheet(ws_transport, plan, header_font, header_fill, thin_border)

        # Sheet 5: Budget
        ws_budget = wb.create_sheet("Budget")
        self._create_budget_sheet(ws_budget, plan, header_font, header_fill, thin_border)

        # Sheet 6: Expert Tips
        ws_tips = wb.create_sheet("Expert Tips")
        self._create_tips_sheet(ws_tips, plan, header_font, header_fill, thin_border)

        # Sheet 7: Weather Forecast (if available)
        if trip_data and trip_data.get("weather"):
            ws_weather = wb.create_sheet("Weather")
            self._create_weather_sheet(ws_weather, trip_data["weather"], header_font, header_fill, thin_border)

        # Sheet 8: Dining & Restaurants (if available)
        if trip_data and trip_data.get("dining"):
            ws_dining = wb.create_sheet("Dining")
            self._create_dining_sheet(ws_dining, trip_data["dining"], header_font, header_fill, thin_border)

        # Sheet 9: Raw Recommendation (for reference)
        ws_raw = wb.create_sheet("Full Recommendation")
        self._create_raw_sheet(ws_raw, plan, title_font)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_overview_sheet(self, ws, plan: TravelPlan, header_font, header_fill, title_font, border):
        """Create the Overview sheet - Professional travel agency style."""
        from openpyxl.styles import Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter

        # Title with branding
        ws['A1'] = "TRAVEL ITINERARY"
        ws['A1'].font = Font(bold=True, size=24, color="2196F3")
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 40

        # Subtitle
        ws['A2'] = plan.overview.destination or "Your Trip"
        ws['A2'].font = Font(bold=True, size=16)
        ws.merge_cells('A2:D2')

        # Trip Summary Box
        ws['A4'] = "TRIP SUMMARY"
        ws['A4'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A4'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        ws.merge_cells('A4:D4')

        # Trip details in a professional format
        overview_data = [
            ("Destination", plan.overview.destination or "Not specified", "Trip Type", plan.overview.trip_type or "Leisure"),
            ("Duration", f"{plan.overview.duration_days} days" if plan.overview.duration_days else "TBD", "Travelers", str(plan.overview.travelers) if plan.overview.travelers else "TBD"),
            ("Total Budget", f"${plan.overview.total_budget:,.2f}" if plan.overview.total_budget else "TBD", "Currency", "USD"),
            ("Document Created", datetime.now().strftime("%B %d, %Y"), "Document ID", f"TP-{datetime.now().strftime('%Y%m%d%H%M')}"),
        ]

        row = 5
        for label1, value1, label2, value2 in overview_data:
            ws.cell(row=row, column=1, value=label1).font = Font(bold=True, size=10)
            ws.cell(row=row, column=2, value=value1).font = Font(size=10)
            ws.cell(row=row, column=3, value=label2).font = Font(bold=True, size=10)
            ws.cell(row=row, column=4, value=value2).font = Font(size=10)
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Quick Reference Section
        ws['A11'] = "QUICK REFERENCE"
        ws['A11'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A11'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells('A11:D11')

        quick_ref = [
            ("Emergency Numbers", "Local: 112 | US Embassy: Check destination"),
            ("Time Zone", "Check destination time zone"),
            ("Voltage", "Check destination electrical standards"),
            ("Language", "Check local language - download offline translator"),
        ]

        row = 12
        for label, value in quick_ref:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
            ws.merge_cells(f'B{row}:D{row}')
            ws.cell(row=row, column=2, value=value).font = Font(size=10)
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Checklist Section
        ws['A18'] = "PRE-TRIP CHECKLIST"
        ws['A18'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A18'].fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells('A18:D18')

        checklist = [
            "[ ] Passport valid for 6+ months",
            "[ ] Visa requirements checked",
            "[ ] Travel insurance purchased",
            "[ ] Accommodations confirmed",
            "[ ] Flights confirmed",
            "[ ] Credit cards notified of travel",
            "[ ] Copies of important documents",
            "[ ] Medications packed",
        ]

        row = 19
        for item in checklist:
            ws.cell(row=row, column=1, value=item).font = Font(size=10)
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30

    def _create_itinerary_sheet(self, ws, plan: TravelPlan, header_font, header_fill, border):
        """Create the Itinerary sheet - Professional travel agency style with booking tracking."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "DAILY ITINERARY"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:J1')
        ws.row_dimensions[1].height = 30

        # Instructions
        ws['A2'] = "Track your daily activities, bookings, and confirmations. Update status as you book."
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:J2')

        # Headers - Professional columns for tracking
        headers = [
            "Day", "Date", "Time Block", "Activity/Event", "Location/Address",
            "Confirmation #", "Contact", "Est. Cost", "Status", "Notes"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row = 5
        days = plan.overview.duration_days or 7

        if plan.itinerary:
            for activity in plan.itinerary:
                ws.cell(row=row, column=1, value=activity.day).border = border
                ws.cell(row=row, column=2, value=activity.date or "").border = border
                ws.cell(row=row, column=3, value=activity.time or "Morning").border = border
                cell = ws.cell(row=row, column=4, value=activity.activity[:80])
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=5, value=activity.location or "").border = border
                ws.cell(row=row, column=6, value="").border = border  # Confirmation #
                ws.cell(row=row, column=7, value="").border = border  # Contact
                ws.cell(row=row, column=8, value=f"${activity.cost:.2f}" if activity.cost else "").border = border
                ws.cell(row=row, column=9, value="Planned").border = border
                ws.cell(row=row, column=10, value=activity.notes or "").border = border
                row += 1
        else:
            # Create empty template rows for each day with time blocks
            time_blocks = ["Morning", "Afternoon", "Evening"]
            for day in range(1, days + 1):
                # Day header row
                ws.cell(row=row, column=1, value=f"Day {day}")
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                for col in range(1, 11):
                    ws.cell(row=row, column=col).border = border
                ws.merge_cells(f'A{row}:J{row}')
                row += 1

                # Time block rows
                for time_block in time_blocks:
                    ws.cell(row=row, column=1, value="").border = border
                    ws.cell(row=row, column=2, value="").border = border  # Date
                    ws.cell(row=row, column=3, value=time_block).border = border
                    ws.cell(row=row, column=4, value="").border = border  # Activity
                    ws.cell(row=row, column=5, value="").border = border  # Location
                    ws.cell(row=row, column=6, value="").border = border  # Confirmation
                    ws.cell(row=row, column=7, value="").border = border  # Contact
                    ws.cell(row=row, column=8, value="").border = border  # Cost
                    ws.cell(row=row, column=9, value="Planned").border = border  # Status
                    ws.cell(row=row, column=10, value="").border = border  # Notes
                    row += 1

        # Status Legend
        row += 2
        ws.cell(row=row, column=1, value="STATUS LEGEND:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        statuses = [
            ("Planned", "E3F2FD", "Activity planned, not yet booked"),
            ("Booked", "C8E6C9", "Reservation confirmed"),
            ("Pending", "FFF9C4", "Awaiting confirmation"),
            ("Cancelled", "FFCDD2", "Cancelled or changed"),
            ("Completed", "B2DFDB", "Activity completed"),
        ]
        for status, color, desc in statuses:
            ws.cell(row=row, column=1, value=status).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=2, value=desc)
            row += 1

        # Column widths
        widths = [8, 12, 12, 40, 30, 15, 15, 12, 12, 25]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_accommodations_sheet(self, ws, plan: TravelPlan, header_font, header_fill, border):
        """Create the Accommodations sheet - Professional hotel booking tracker."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "ACCOMMODATION TRACKER"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 30

        # Summary section
        ws['A3'] = "BOOKING SUMMARY"
        ws['A3'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        ws.merge_cells('A3:L3')

        nights = plan.overview.duration_days or 7
        summary_data = [
            ("Total Nights", str(nights), "Check-in Date", ""),
            ("Total Accommodations", "0", "Check-out Date", ""),
            ("Est. Accommodation Cost", "$0.00", "Total Guests", str(plan.overview.travelers or 2)),
        ]

        row = 4
        for label1, value1, label2, value2 in summary_data:
            ws.cell(row=row, column=1, value=label1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value1)
            ws.cell(row=row, column=4, value=label2).font = Font(bold=True)
            ws.cell(row=row, column=5, value=value2)
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Main booking table header
        row += 1
        ws.cell(row=row, column=1, value="ACCOMMODATION DETAILS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells(f'A{row}:L{row}')

        row += 1
        headers = [
            "Property Name", "Type", "Address/Location", "Check-in", "Check-out",
            "Nights", "Rate/Night", "Total Cost", "Confirmation #",
            "Contact Phone", "Booking Status", "Payment Status"
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row += 1
        start_data_row = row

        if plan.accommodations:
            for acc in plan.accommodations:
                ws.cell(row=row, column=1, value=acc.name[:40]).border = border
                ws.cell(row=row, column=2, value=acc.type or "Hotel").border = border
                ws.cell(row=row, column=3, value=acc.location or "").border = border
                ws.cell(row=row, column=4, value=acc.check_in or "").border = border
                ws.cell(row=row, column=5, value=acc.check_out or "").border = border
                ws.cell(row=row, column=6, value="").border = border  # Nights
                ws.cell(row=row, column=7, value=f"${acc.price_per_night:.2f}" if acc.price_per_night else "").border = border
                ws.cell(row=row, column=8, value=f"${acc.total_price:.2f}" if acc.total_price else "").border = border
                ws.cell(row=row, column=9, value="").border = border  # Confirmation
                ws.cell(row=row, column=10, value="").border = border  # Contact
                ws.cell(row=row, column=11, value="Researching").border = border
                ws.cell(row=row, column=12, value="Not Paid").border = border
                row += 1
        else:
            # Empty template rows
            for i in range(3):
                for col in range(1, 13):
                    ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=11, value="Researching").border = border
                ws.cell(row=row, column=12, value="Not Paid").border = border
                row += 1

        # Status legends
        row += 2
        ws.cell(row=row, column=1, value="BOOKING STATUS:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=4, value="PAYMENT STATUS:")
        ws.cell(row=row, column=4).font = Font(bold=True)
        row += 1

        booking_statuses = [
            ("Researching", "E3F2FD"),
            ("Requested", "FFF9C4"),
            ("Confirmed", "C8E6C9"),
            ("Cancelled", "FFCDD2"),
        ]
        payment_statuses = [
            ("Not Paid", "FFCDD2"),
            ("Deposit Paid", "FFF9C4"),
            ("Fully Paid", "C8E6C9"),
            ("Refunded", "E1BEE7"),
        ]

        for i, ((b_status, b_color), (p_status, p_color)) in enumerate(zip(booking_statuses, payment_statuses)):
            ws.cell(row=row + i, column=1, value=b_status).fill = PatternFill(start_color=b_color, end_color=b_color, fill_type="solid")
            ws.cell(row=row + i, column=4, value=p_status).fill = PatternFill(start_color=p_color, end_color=p_color, fill_type="solid")

        # Important reminders section
        row += 6
        ws.cell(row=row, column=1, value="IMPORTANT REMINDERS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells(f'A{row}:L{row}')

        reminders = [
            "[ ] Confirm check-in time and early check-in availability",
            "[ ] Note cancellation policy and deadline",
            "[ ] Save confirmation emails/documents",
            "[ ] Check room preferences (bed type, floor, view)",
            "[ ] Verify amenities (WiFi, parking, breakfast)",
            "[ ] Register for loyalty programs if applicable",
        ]
        row += 1
        for reminder in reminders:
            ws.cell(row=row, column=1, value=reminder)
            ws.merge_cells(f'A{row}:L{row}')
            row += 1

        # Column widths
        widths = [25, 10, 30, 12, 12, 8, 12, 12, 15, 15, 12, 12]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_transportation_sheet(self, ws, plan: TravelPlan, header_font, header_fill, border):
        """Create the Transportation sheet - Professional flight & transport tracker."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "TRANSPORTATION TRACKER"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:N1')
        ws.row_dimensions[1].height = 30

        # Flight Summary Section
        ws['A3'] = "FLIGHTS & TRANSPORTATION SUMMARY"
        ws['A3'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        ws.merge_cells('A3:N3')

        summary_data = [
            ("Total Segments", "0", "Total Cost", "$0.00"),
            ("Departure City", "", "Return City", ""),
            ("Outbound Date", "", "Return Date", ""),
        ]

        row = 4
        for label1, value1, label2, value2 in summary_data:
            ws.cell(row=row, column=1, value=label1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value1)
            ws.cell(row=row, column=4, value=label2).font = Font(bold=True)
            ws.cell(row=row, column=5, value=value2)
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Flights section
        row += 1
        ws.cell(row=row, column=1, value="FLIGHT DETAILS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}')

        row += 1
        flight_headers = [
            "Segment", "Date", "Flight #", "Airline", "From", "To",
            "Depart", "Arrive", "Duration", "Terminal", "Seat",
            "PNR/Conf #", "Check-in Status", "Baggage"
        ]

        for col, header in enumerate(flight_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row += 1

        if plan.transportation:
            segment = 1
            for trans in plan.transportation:
                if trans.type.lower() == "flight":
                    ws.cell(row=row, column=1, value=f"Leg {segment}").border = border
                    ws.cell(row=row, column=2, value=trans.date or "").border = border
                    ws.cell(row=row, column=3, value="").border = border  # Flight #
                    ws.cell(row=row, column=4, value=trans.carrier or "").border = border
                    ws.cell(row=row, column=5, value=trans.from_location or "").border = border
                    ws.cell(row=row, column=6, value=trans.to_location or "").border = border
                    ws.cell(row=row, column=7, value=trans.departure_time or "").border = border
                    ws.cell(row=row, column=8, value=trans.arrival_time or "").border = border
                    ws.cell(row=row, column=9, value=trans.duration or "").border = border
                    ws.cell(row=row, column=10, value="").border = border  # Terminal
                    ws.cell(row=row, column=11, value="").border = border  # Seat
                    ws.cell(row=row, column=12, value="").border = border  # PNR
                    ws.cell(row=row, column=13, value="Not Checked In").border = border
                    ws.cell(row=row, column=14, value="").border = border  # Baggage
                    segment += 1
                    row += 1
        else:
            # Empty template for outbound and return flights
            for leg in ["Outbound", "Return"]:
                ws.cell(row=row, column=1, value=leg).border = border
                for col in range(2, 15):
                    ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=13, value="Not Checked In").border = border
                row += 1

        # Ground Transportation Section
        row += 2
        ws.cell(row=row, column=1, value="GROUND TRANSPORTATION")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}')

        row += 1
        ground_headers = [
            "Type", "Date", "Time", "From", "To", "Company/Driver",
            "Confirmation #", "Contact", "Cost", "Status", "Notes", "", "", ""
        ]

        for col, header in enumerate(ground_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row += 1
        # Template rows for ground transport
        ground_types = ["Airport Transfer (Arrival)", "Airport Transfer (Departure)", "Car Rental", "Other"]
        for g_type in ground_types:
            ws.cell(row=row, column=1, value=g_type).border = border
            for col in range(2, 15):
                ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=10, value="Pending").border = border
            row += 1

        # Check-in reminder section
        row += 2
        ws.cell(row=row, column=1, value="PRE-FLIGHT CHECKLIST")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}')

        checklist = [
            "[ ] Online check-in (24 hrs before departure)",
            "[ ] Boarding passes downloaded/printed",
            "[ ] Passport/ID ready (valid 6+ months)",
            "[ ] Baggage packed within weight limits",
            "[ ] Liquids in clear bag (carry-on)",
            "[ ] Arrive at airport 2-3 hrs before international flights",
            "[ ] TSA PreCheck / Global Entry if applicable",
            "[ ] Seat selection confirmed",
        ]

        row += 1
        for item in checklist:
            ws.cell(row=row, column=1, value=item)
            ws.merge_cells(f'A{row}:N{row}')
            row += 1

        # Status legends
        row += 2
        ws.cell(row=row, column=1, value="CHECK-IN STATUS:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        statuses = [
            ("Not Checked In", "FFCDD2"),
            ("Checked In", "C8E6C9"),
            ("Boarding Pass Ready", "B2DFDB"),
            ("Cancelled", "E0E0E0"),
        ]

        for status, color in statuses:
            ws.cell(row=row, column=1, value=status).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1

        # Column widths
        widths = [12, 12, 10, 15, 15, 15, 10, 10, 10, 10, 10, 15, 14, 12]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_budget_sheet(self, ws, plan: TravelPlan, header_font, header_fill, border):
        """Create the Budget sheet - Professional travel agency style with tracking."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "TRIP BUDGET TRACKER"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:F1')

        # Budget Summary Box
        ws['A3'] = "BUDGET SUMMARY"
        ws['A3'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        ws.merge_cells('A3:F3')

        total_budget = plan.overview.total_budget or 5000
        travelers = plan.overview.travelers or 2
        days = plan.overview.duration_days or 7

        summary_data = [
            ("Total Budget", f"${total_budget:,.2f}", "Per Person", f"${total_budget/travelers:,.2f}"),
            ("Trip Duration", f"{days} days", "Daily Budget", f"${total_budget/days:,.2f}"),
        ]

        row = 4
        for label1, value1, label2, value2 in summary_data:
            ws.cell(row=row, column=1, value=label1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value1)
            ws.cell(row=row, column=4, value=label2).font = Font(bold=True)
            ws.cell(row=row, column=5, value=value2)
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Budget Breakdown Header
        ws['A8'] = "BUDGET BREAKDOWN"
        ws['A8'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A8'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells('A8:F8')

        # Headers for tracking
        headers = ["Category", "Estimated", "Actual", "Difference", "Status", "Notes"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Pre-defined budget categories with typical percentages
        budget_categories = [
            ("Flights/Transportation", total_budget * 0.30, "Round-trip airfare + ground transport"),
            ("Accommodation", total_budget * 0.25, f"{days} nights"),
            ("Food & Dining", total_budget * 0.20, f"${total_budget * 0.20 / days:.0f}/day"),
            ("Activities & Tours", total_budget * 0.15, "Attractions, tours, experiences"),
            ("Shopping & Souvenirs", total_budget * 0.05, "Gifts, souvenirs"),
            ("Emergency Fund", total_budget * 0.05, "Unexpected expenses"),
        ]

        row = 10
        total_estimated = 0
        for category, estimated, notes in budget_categories:
            ws.cell(row=row, column=1, value=category).border = border
            ws.cell(row=row, column=2, value=f"${estimated:,.2f}").border = border
            ws.cell(row=row, column=3, value="$0.00").border = border  # Actual - user fills in
            ws.cell(row=row, column=4, value=f"=${chr(66)}{row}-{chr(67)}{row}").border = border  # Formula
            ws.cell(row=row, column=5, value="Planned").border = border
            ws.cell(row=row, column=6, value=notes).border = border
            total_estimated += estimated
            row += 1

        # Total row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"${total_estimated:,.2f}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"=SUM(C10:C{row-1})").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").font = Font(bold=True)
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

        # Expense Log Section
        row += 3
        ws.cell(row=row, column=1, value="EXPENSE LOG")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells(f'A{row}:F{row}')

        row += 1
        log_headers = ["Date", "Category", "Description", "Amount", "Payment Method", "Receipt"]
        for col, header in enumerate(log_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Empty rows for logging expenses
        for i in range(10):
            row += 1
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border

        # Column widths
        widths = [25, 15, 15, 15, 15, 35]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_tips_sheet(self, ws, plan: TravelPlan, header_font, header_fill, border):
        """Create the Expert Tips sheet - Professional action-oriented tips tracker."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "EXPERT RECOMMENDATIONS & ACTION ITEMS"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 30

        # Instructions
        ws['A2'] = "Track expert advice and action items. Check off completed tasks before your trip."
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:G2')

        # Summary by category
        ws['A4'] = "TIPS BY CATEGORY"
        ws['A4'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A4'].fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        ws.merge_cells('A4:G4')

        # Categories with colors
        categories = [
            ("Budget & Money", "4CAF50", "Currency, ATMs, budgeting tips"),
            ("Safety & Health", "F44336", "Insurance, vaccines, emergency contacts"),
            ("Culture & Etiquette", "9C27B0", "Local customs, dress code, language"),
            ("Food & Dining", "FF9800", "Must-try dishes, dietary notes"),
            ("Activities & Tours", "2196F3", "Booking recommendations, timing"),
            ("Logistics", "607D8B", "Transportation, packing, timing"),
            ("Weather & Packing", "00BCD4", "What to pack, seasonal tips"),
            ("Accommodation", "795548", "Check-in tips, location advice"),
        ]

        row = 5
        for cat_name, color, description in categories:
            ws.cell(row=row, column=1, value=cat_name).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=2, value=description)
            ws.merge_cells(f'B{row}:G{row}')
            row += 1

        # Action items header
        row += 1
        ws.cell(row=row, column=1, value="ACTION ITEMS & RECOMMENDATIONS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells(f'A{row}:G{row}')

        row += 1
        headers = ["Category", "Expert Source", "Recommendation/Action", "Priority", "Status", "Due Date", "Notes"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row += 1

        if plan.expert_tips:
            for tip in plan.expert_tips:
                ws.cell(row=row, column=1, value=tip.category).border = border
                ws.cell(row=row, column=2, value=tip.expert).border = border
                cell = ws.cell(row=row, column=3, value=tip.tip[:150])
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

                # Priority with color coding
                priority_cell = ws.cell(row=row, column=4, value=tip.priority)
                priority_cell.border = border
                if tip.priority == "High":
                    priority_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                elif tip.priority == "Medium":
                    priority_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                else:
                    priority_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                ws.cell(row=row, column=5, value="Not Started").border = border
                ws.cell(row=row, column=6, value="").border = border  # Due date
                ws.cell(row=row, column=7, value="").border = border  # Notes
                row += 1
        else:
            # Add template action items common to all trips
            common_actions = [
                ("Safety", "Safety Expert", "Purchase travel insurance", "High"),
                ("Safety", "Safety Expert", "Check travel advisories for destination", "High"),
                ("Budget", "Budget Advisor", "Notify bank of travel dates", "High"),
                ("Budget", "Budget Advisor", "Research currency exchange rates", "Medium"),
                ("Logistics", "Logistics Planner", "Arrange airport transportation", "High"),
                ("Culture", "Culture Guide", "Learn basic local phrases", "Medium"),
                ("Weather", "Weather Analyst", "Check weather forecast closer to trip", "Medium"),
                ("Food", "Food Expert", "Research must-try local dishes", "Low"),
            ]

            for category, expert, action, priority in common_actions:
                ws.cell(row=row, column=1, value=category).border = border
                ws.cell(row=row, column=2, value=expert).border = border
                ws.cell(row=row, column=3, value=action).border = border

                priority_cell = ws.cell(row=row, column=4, value=priority)
                priority_cell.border = border
                if priority == "High":
                    priority_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                elif priority == "Medium":
                    priority_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                else:
                    priority_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

                ws.cell(row=row, column=5, value="Not Started").border = border
                ws.cell(row=row, column=6, value="").border = border
                ws.cell(row=row, column=7, value="").border = border
                row += 1

        # Status legend
        row += 2
        ws.cell(row=row, column=1, value="STATUS OPTIONS:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        statuses = [
            ("Not Started", "E0E0E0", "Task not yet begun"),
            ("In Progress", "FFF9C4", "Currently working on"),
            ("Completed", "C8E6C9", "Done and verified"),
            ("N/A", "E1BEE7", "Not applicable to this trip"),
        ]

        for status, color, desc in statuses:
            ws.cell(row=row, column=1, value=status).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=2, value=desc)
            row += 1

        # Priority legend
        row += 1
        ws.cell(row=row, column=1, value="PRIORITY LEVELS:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        priorities = [
            ("High", "FFCDD2", "Complete before booking/travel"),
            ("Medium", "FFF9C4", "Important but flexible timing"),
            ("Low", "C8E6C9", "Nice to have, do if time permits"),
        ]

        for priority, color, desc in priorities:
            ws.cell(row=row, column=1, value=priority).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=2, value=desc)
            row += 1

        # Column widths
        widths = [15, 20, 50, 10, 12, 12, 25]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_raw_sheet(self, ws, plan: TravelPlan, title_font):
        """Create the raw recommendation sheet for reference."""
        from openpyxl.styles import Alignment

        ws['A1'] = "Full Travel Recommendation"
        ws['A1'].font = title_font

        ws['A3'] = "Original Question:"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A4:H8')
        ws['A4'] = plan.raw_question or "No question recorded"
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')

        ws['A10'] = "Synthesized Recommendation:"
        ws['A10'].font = Font(bold=True)
        ws.merge_cells('A11:H50')
        ws['A11'] = plan.raw_recommendation or "No recommendation recorded"
        ws['A11'].alignment = Alignment(wrap_text=True, vertical='top')

        # Column width
        ws.column_dimensions['A'].width = 100

    def _create_weather_sheet(self, ws, weather_data: str, header_font, header_fill, border):
        """Create the Weather Forecast sheet - Professional packing guide."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "WEATHER FORECAST & PACKING GUIDE"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        # Weather Forecast Section
        ws['A3'] = "DAILY WEATHER FORECAST"
        ws['A3'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A3'].fill = PatternFill(start_color="00BCD4", end_color="00BCD4", fill_type="solid")
        ws.merge_cells('A3:H3')

        row = 4
        headers = ["Date", "Condition", "High", "Low", "Humidity", "Rain %", "Wind", "Pack For"]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        if weather_data:
            # Parse weather data from formatted string
            lines = weather_data.split('\n')

            for line in lines:
                # Skip header/separator lines
                if '|' not in line or '---' in line or 'Date' in line:
                    continue

                # Parse table row: | Date | Weather | Temp | Humidity | Rain % | Wind |
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if len(parts) >= 5:
                    ws.cell(row=row, column=1, value=parts[0]).border = border  # Date
                    condition = parts[1] if len(parts) > 1 else ""
                    ws.cell(row=row, column=2, value=condition).border = border  # Condition

                    # Parse temp range (e.g., "45°-58°")
                    temp_str = parts[2] if len(parts) > 2 else ""
                    if '-' in temp_str:
                        temps = temp_str.replace('°', '').split('-')
                        ws.cell(row=row, column=3, value=temps[1] if len(temps) > 1 else "").border = border  # High
                        ws.cell(row=row, column=4, value=temps[0]).border = border  # Low
                    else:
                        ws.cell(row=row, column=3, value=temp_str).border = border
                        ws.cell(row=row, column=4, value="").border = border

                    ws.cell(row=row, column=5, value=parts[3] if len(parts) > 3 else "").border = border  # Humidity
                    rain_pct = parts[4] if len(parts) > 4 else ""
                    ws.cell(row=row, column=6, value=rain_pct).border = border  # Rain %
                    ws.cell(row=row, column=7, value=parts[5] if len(parts) > 5 else "").border = border  # Wind

                    # Auto-suggest packing based on conditions
                    pack_suggestion = ""
                    if "rain" in condition.lower() or (rain_pct and int(rain_pct.replace('%', '')) > 30):
                        pack_suggestion = "Umbrella/Rain jacket"
                    elif "cloud" in condition.lower():
                        pack_suggestion = "Light layers"
                    elif "sun" in condition.lower() or "clear" in condition.lower():
                        pack_suggestion = "Sunscreen, sunglasses"
                    ws.cell(row=row, column=8, value=pack_suggestion).border = border

                    row += 1
        else:
            ws.cell(row=row, column=1, value="Weather data will be available closer to your trip date.")
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

        # Packing Checklist Section
        row += 2
        ws.cell(row=row, column=1, value="WEATHER-BASED PACKING CHECKLIST")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells(f'A{row}:H{row}')

        row += 1
        packing_categories = [
            ("CLOTHING", ["[ ] Light layers", "[ ] Rain jacket/umbrella", "[ ] Comfortable walking shoes", "[ ] Dressy outfit for dinners"]),
            ("SUN PROTECTION", ["[ ] Sunscreen (SPF 30+)", "[ ] Sunglasses", "[ ] Hat/Cap", "[ ] Lip balm with SPF"]),
            ("COLD WEATHER", ["[ ] Warm jacket", "[ ] Scarf/Gloves", "[ ] Thermal underlayers", "[ ] Warm socks"]),
            ("RAINY WEATHER", ["[ ] Waterproof jacket", "[ ] Umbrella", "[ ] Waterproof bag cover", "[ ] Quick-dry clothing"]),
        ]

        for category, items in packing_categories:
            ws.cell(row=row, column=1, value=category).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

            col = 1
            for item in items:
                ws.cell(row=row, column=col, value=item)
                col += 2
                if col > 7:
                    col = 1
                    row += 1
            if col != 1:
                row += 1

        # Weather Tips Section
        row += 1
        ws.cell(row=row, column=1, value="WEATHER TIPS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells(f'A{row}:H{row}')

        tips = [
            "Check weather forecast again 2-3 days before departure",
            "Weather apps: Weather.com, AccuWeather, or destination's local weather service",
            "Pack versatile layers that can be mixed and matched",
            "Consider the indoor/outdoor ratio of your activities",
            "Research seasonal events (monsoons, festivals, peak seasons)",
        ]

        row += 1
        for tip in tips:
            ws.cell(row=row, column=1, value=f"• {tip}")
            ws.merge_cells(f'A{row}:H{row}')
            row += 1

        # Column widths
        widths = [15, 20, 10, 10, 12, 10, 12, 25]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def _create_dining_sheet(self, ws, dining_data: str, header_font, header_fill, border):
        """Create the Dining & Restaurants sheet - Professional restaurant tracker."""
        from openpyxl.styles import Alignment, Font, PatternFill

        # Title
        ws['A1'] = "DINING & RESTAURANT PLANNER"
        ws['A1'].font = Font(bold=True, size=16, color="2196F3")
        ws.merge_cells('A1:K1')
        ws.row_dimensions[1].height = 30

        # Instructions
        ws['A2'] = "Track restaurant recommendations, reservations, and dining experiences."
        ws['A2'].font = Font(italic=True, size=10, color="666666")
        ws.merge_cells('A2:K2')

        # Reservation Tracker Section
        ws['A4'] = "RESTAURANT RESERVATIONS"
        ws['A4'].font = Font(bold=True, size=11, color="FFFFFF")
        ws['A4'].fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        ws.merge_cells('A4:K4')

        row = 5
        reservation_headers = [
            "Restaurant Name", "Cuisine", "Date", "Time", "Party Size",
            "Confirmation #", "Contact/Phone", "Address", "Price Range",
            "Status", "Notes"
        ]

        for col, header in enumerate(reservation_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        row += 1
        # Empty template rows for reservations
        for i in range(5):
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=10, value="Not Booked").border = border
            row += 1

        # Recommended Restaurants Section
        row += 1
        ws.cell(row=row, column=1, value="RECOMMENDED RESTAURANTS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        ws.merge_cells(f'A{row}:K{row}')

        row += 1
        if dining_data:
            # Add the raw recommendations in a formatted section
            ws.cell(row=row, column=1, value="AI-Generated Recommendations:")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # Split dining data into manageable chunks
            lines = dining_data.split('\n')
            for line in lines[:40]:  # Limit to 40 lines
                if line.strip():
                    ws.cell(row=row, column=1, value=line.strip()[:150])
                    ws.merge_cells(f'A{row}:K{row}')
                    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
                    row += 1
        else:
            ws.cell(row=row, column=1, value="No AI recommendations available. Add restaurants manually above.")
            ws.merge_cells(f'A{row}:K{row}')
            row += 1

        # Dining Budget Section
        row += 2
        ws.cell(row=row, column=1, value="DINING BUDGET")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        ws.merge_cells(f'A{row}:K{row}')

        row += 1
        budget_headers = ["Meal Type", "Daily Budget", "# Days", "Subtotal", "Notes"]
        for col, header in enumerate(budget_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        row += 1
        meal_types = [
            ("Breakfast", "$15-25"),
            ("Lunch", "$20-35"),
            ("Dinner", "$40-75"),
            ("Snacks/Coffee", "$10-20"),
            ("Special Dining", "$100+"),
        ]

        for meal, budget_range in meal_types:
            ws.cell(row=row, column=1, value=meal).border = border
            ws.cell(row=row, column=2, value=budget_range).border = border
            ws.cell(row=row, column=3, value="").border = border  # Days
            ws.cell(row=row, column=4, value="").border = border  # Subtotal
            ws.cell(row=row, column=5, value="").border = border  # Notes
            row += 1

        # Status Legend
        row += 2
        ws.cell(row=row, column=1, value="RESERVATION STATUS:")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        statuses = [
            ("Not Booked", "E0E0E0", "Haven't made reservation yet"),
            ("Requested", "FFF9C4", "Reservation request sent"),
            ("Confirmed", "C8E6C9", "Reservation confirmed"),
            ("Waitlist", "FFCDD2", "On waitlist"),
            ("Cancelled", "E1BEE7", "Reservation cancelled"),
        ]

        for status, color, desc in statuses:
            ws.cell(row=row, column=1, value=status).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=2, value=desc)
            row += 1

        # Dietary Notes Section
        row += 2
        ws.cell(row=row, column=1, value="DIETARY NOTES & PREFERENCES")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="607D8B", end_color="607D8B", fill_type="solid")
        ws.merge_cells(f'A{row}:K{row}')

        row += 1
        dietary_items = [
            "Allergies:",
            "Dietary Restrictions:",
            "Food Preferences:",
            "Must-Try Local Dishes:",
        ]

        for item in dietary_items:
            ws.cell(row=row, column=1, value=item).font = Font(bold=True)
            ws.merge_cells(f'B{row}:K{row}')
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = border
            row += 1

        # Column widths
        widths = [20, 15, 12, 10, 10, 15, 15, 25, 12, 12, 20]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + col)].width = width

    def export_with_trip_data(
        self,
        trip_config: Dict,
        trip_data: Dict,
        expert_responses: Dict[str, Dict[str, Any]]
    ) -> BytesIO:
        """
        Export travel plan with real-time trip data.

        Args:
            trip_config: Trip configuration (destination, dates, etc.)
            trip_data: Real-time data (weather, flights, hotels, dining)
            expert_responses: Dict of expert responses

        Returns:
            BytesIO buffer containing Excel workbook
        """
        # Build question from config
        destination = trip_config.get("destination", "")
        question = f"Trip to {destination}, {trip_config.get('travelers', '2 adults')}, budget ${trip_config.get('budget', 5000)}"
        recommendation = trip_data.get("summary", "") if trip_data else ""

        # Use the standard export method
        return self.export_to_excel(
            question=question,
            recommendation=recommendation,
            expert_responses=expert_responses
        )


def export_travel_plan_to_excel(
    question: str,
    recommendation: str,
    expert_responses: Dict[str, Dict[str, Any]],
    trip_data: Optional[Dict] = None
) -> BytesIO:
    """
    Convenience function to export travel plan to Excel.

    Args:
        question: Original travel question
        recommendation: Final synthesized recommendation
        expert_responses: Dict of expert responses
        trip_data: Optional real-time trip data (weather, flights, hotels, dining)

    Returns:
        BytesIO buffer containing Excel workbook
    """
    service = ExcelExportService()
    return service.export_to_excel(question, recommendation, expert_responses, trip_data=trip_data)
